import requests
import json
import time
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

API_URL = "http://localhost:5000/api/search"
LOG_FILE = "chatbot_log.xlsx"

# ==========================================================
# 🔹 Fungsi Cetak & Format
# ==========================================================
def print_header(title):
    print("=" * 70)
    print(title)
    print("=" * 70)

def print_step(title, content=None):
    print(f"\n🧩 {title}")
    print("-" * 70)
    if content:
        if isinstance(content, dict):
            for k, v in content.items():
                print(f"{k:20}: {v}")
        else:
            print(content)
    print("-" * 70)

def print_timing(t):
    print(f"⏱️  AI Filter     : {t.get('ai_domain_sec', '-')}s")
    print(f"🤖 AI Relevance  : {t.get('ai_relevance_sec', '-')}s")
    print(f"🔢 Embedding     : {t.get('embedding_sec', '-')}s")
    print(f"🗃️  Qdrant        : {t.get('qdrant_sec', '-')}s")
    print(f"⚡ Total Time    : {t.get('total_sec', '-')}s")

def print_candidates(results):
    print(f"\n🔍 Candidate Results ({len(results)} ditemukan)")
    print("-" * 70)
    for i, r in enumerate(results, start=1):
        print(f"[{i}] {r['question']}")
        print(f"    Dense: {r['dense_score']:.3f} | Overlap: {r['overlap_score']:.3f} | Final: {r['final_score']:.3f} | Note: {r['note']}")
    print("-" * 70)

# ==========================================================
# 🔹 Logging ke Excel (Tanpa Pandas)
# ==========================================================
def log_to_excel(entry):
    columns = [
        "Timestamp", "Status", "Original Question", "Final Question",
        "Question Sent to RAG", "AI Reformulated Question",
        "Category", "Dense Top", "Final Score Top",
        "AI Reason", "AI Reformulated",
        "Total Candidates", "Accepted", "Rejected",
        "AI Filter (s)", "AI Relevance (s)", "Embedding (s)",
        "Qdrant (s)", "Total Time (s)", "Total Processing Time (s)"
    ]

    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(columns)
        wb.save(LOG_FILE)

    wb = load_workbook(LOG_FILE)
    ws = wb.active
    ws.append([
        entry.get("Timestamp", ""),
        entry.get("Status", ""),
        entry.get("Original Question", ""),
        entry.get("Final Question", ""),
        entry.get("Question Sent to RAG", ""),
        entry.get("AI Reformulated Question", ""),
        entry.get("Category", ""),
        entry.get("Dense Top", ""),
        entry.get("Final Score Top", ""),
        entry.get("AI Reason", ""),
        entry.get("AI Reformulated", ""),
        entry.get("Total Candidates", ""),
        entry.get("Accepted", ""),
        entry.get("Rejected", ""),
        entry.get("AI Filter (s)", ""),
        entry.get("AI Relevance (s)", ""),
        entry.get("Embedding (s)", ""),
        entry.get("Qdrant (s)", ""),
        entry.get("Total Time (s)", ""),
        entry.get("Total Processing Time (s)", "")
    ])
    wb.save(LOG_FILE)

# ==========================================================
# 🔹 Fungsi Utama
# ==========================================================
def main():
    print("🤖 Chatbot Debugger (ketik 'exit' untuk keluar)\n")

    while True:
        question = input("Anda: ").strip()
        if question.lower() in ["exit", "quit"]:
            print("👋 Keluar...")
            break

        payload = {"question": question}
        try:
            t_start = time.time()
            response = requests.post(API_URL, json=payload, timeout=120)
            total_elapsed = round(time.time() - t_start, 2)

            print_header(f"FULL PIPELINE TRACE ({total_elapsed}s)")
            print_step("📥 Input Pertanyaan", question)

            if response.status_code != 200:
                print(f"❌ Server error: {response.status_code}")
                print(response.text)
                continue

            data = response.json()
            status = data.get("status", "").upper()
            timing = data.get("timing", {})
            meta = data.get("data", {}).get("metadata", {})

            # 1️⃣ PRE-FILTER
            print_step("🔍 PRE-PROCESSING (AI Filter / Hard Filter)")
            print(f"AI Filter Time : {timing.get('ai_domain_sec', '-')}s")

            # 2️⃣ EMBEDDING
            print_step("🧠 EMBEDDING VECTOR")
            print("Embedding dilakukan dengan model e5-small-local (Flask API)\n-> hasil vektor tidak ditampilkan demi efisiensi.")

            # 3️⃣ QDRANT
            print_step("🗃️  QDRANT SEARCH")
            print(f"Qdrant Search Time : {timing.get('qdrant_sec', '-')}")
            if status == "SUCCESS":
                print(f"Category Detected  : {meta.get('category', '-')}")
                print(f"Top Dense Score    : {meta.get('dense_score_top', '-')}")
                print(f"Top Final Score    : {meta.get('final_score_top', '-')}")
            else:
                print("Qdrant hasil ditemukan namun tidak cukup relevan.")

            # 4️⃣ AI RELEVANCE
            print_step("🤖 POST-PROCESSING (AI Relevance Check)")
            ai_reason = meta.get("ai_reason", "-")
            ai_reform = meta.get("ai_reformulated", "-")
            print(f"AI Reason       : {ai_reason}")
            print(f"Reformulated Q  : {ai_reform}")

            # 5️⃣ HASIL AKHIR
            print_step("📤 OUTPUT AKHIR")
            accepted = 0
            rejected = 0
            dense_top = meta.get("dense_score_top", 0)
            final_top = meta.get("final_score_top", 0)
            category = meta.get("category", "-")
            final_q = meta.get("final_question", question)
            ai_reform_q = meta.get("ai_reformulated", "-")

            if status == "SUCCESS":
                print(f"STATUS        : SUCCESS")
                print(f"Original Q    : {meta.get('original_question', '-')}")
                print(f"Final Q       : {final_q}")
                print(f"Category      : {category}")
                print(f"AI Reform     : {ai_reform_q}")
                print_candidates(data["data"]["similar_questions"])
                accepted = len(data["data"]["similar_questions"])
            else:
                print("STATUS        : LOW CONFIDENCE")
                print(f"Message       : {data.get('message', '-')}")

            # 6️⃣ TIMING
            print_step("⏱️  WAKTU EKSEKUSI")
            print_timing(timing)
            print(f"🕒 Total Processing Time (client): {total_elapsed}s")

            print("\n" + "=" * 70 + "\n")

            # 7️⃣ LOGGING KE EXCEL
            entry = {
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Status": status,
                "Original Question": question,
                "Final Question": final_q,
                "Question Sent to RAG": final_q if status == "SUCCESS" else "-",
                "AI Reformulated Question": ai_reform_q,
                "Category": category,
                "Dense Top": dense_top,
                "Final Score Top": final_top,
                "AI Reason": ai_reason,
                "AI Reformulated": ai_reform_q,
                "Total Candidates": accepted + rejected,
                "Accepted": accepted,
                "Rejected": rejected,
                "AI Filter (s)": timing.get("ai_domain_sec", 0),
                "AI Relevance (s)": timing.get("ai_relevance_sec", 0),
                "Embedding (s)": timing.get("embedding_sec", 0),
                "Qdrant (s)": timing.get("qdrant_sec", 0),
                "Total Time (s)": timing.get("total_sec", 0),
                "Total Processing Time (s)": total_elapsed
            }

            log_to_excel(entry)
            print(f"📝 Log tersimpan ke {LOG_FILE}\n")

        except Exception as e:
            print(f"⚠️  ERROR: {str(e)}")
            print("=" * 60)
            print()

# ==========================================================
# 🔹 Jalankan Program
# ==========================================================
if __name__ == "__main__":
    main()
